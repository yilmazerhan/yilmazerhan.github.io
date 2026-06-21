"""Export router — CSV/Excel export for work logs, tasks, user activity."""
import csv
import io
import uuid
from datetime import date
from typing import Annotated, Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.database import get_db
from app.models.user import User
from app.models.user_team import user_teams
from app.models.worklog import WorkLog
from app.models.kanban import Task, KanbanColumn
from app.models.task_label import task_label_assignments, TaskLabel
from app.core.dependencies import get_current_user

router = APIRouter(prefix="/export", tags=["export"])

_FORMULA_CHARS = ('=', '+', '-', '@', '\t', '\r')


def _csv_safe(value: str) -> str:
    """Prevent CSV formula injection by prefixing formula-like values."""
    if value and value[0] in _FORMULA_CHARS:
        return "'" + value
    return value


def _csv_response(filename: str, rows: list[list], headers: list[str]) -> StreamingResponse:
    """Return a StreamingResponse with CSV content."""
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(headers)
    writer.writerows(rows)
    buffer.seek(0)
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="text/csv; charset=utf-8-sig",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─── Work Log Export ─────────────────────────────────────────────────────────

@router.get("/worklogs")
async def export_worklogs(
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    user_id: Optional[uuid.UUID] = Query(None),
    format: str = Query("csv", pattern="^(csv|excel)$"),
):
    """Export work logs with optional date range and user filter."""
    from app.models.worklog import WorkLog, WorkType

    q = (
        select(WorkLog)
        .options(
            selectinload(WorkLog.user),
            selectinload(WorkLog.work_type),
        )
    )

    # Role-based scoping
    if current_user.role == "superadmin":
        if user_id:
            q = q.where(WorkLog.user_id == user_id)
    elif current_user.role == "team_manager":
        # Scope to users in the manager's teams (use junction table — authoritative)
        my_team_ids = select(user_teams.c.team_id).where(user_teams.c.user_id == current_user.id)
        team_member_ids = select(user_teams.c.user_id).where(user_teams.c.team_id.in_(my_team_ids))
        q = q.where(WorkLog.user_id.in_(team_member_ids))
        if user_id:
            q = q.where(WorkLog.user_id == user_id)
    else:
        # Regular user sees only own logs
        q = q.where(WorkLog.user_id == current_user.id)

    if date_from:
        q = q.where(WorkLog.log_date >= date_from)
    if date_to:
        q = q.where(WorkLog.log_date <= date_to)

    q = q.order_by(WorkLog.log_date.desc(), WorkLog.created_at.desc())
    result = await db.execute(q)
    logs = list(result.scalars().all())

    headers = ["Date", "User", "Work Type", "Hours", "Description"]
    rows = [
        [
            log.log_date.isoformat() if log.log_date else "",
            _csv_safe(log.user.full_name if log.user else ""),
            _csv_safe(log.work_type.name if log.work_type else ""),
            str(log.duration_hours),
            _csv_safe(log.description or ""),
        ]
        for log in logs
    ]

    if format == "excel":
        return _excel_response("worklogs.xlsx", rows, headers)

    return _csv_response("worklogs.csv", rows, headers)


# ─── Task Export ──────────────────────────────────────────────────────────────

@router.get("/tasks")
async def export_tasks(
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    board_id: Optional[uuid.UUID] = Query(None),
    column_id: Optional[uuid.UUID] = Query(None),
    assignee_id: Optional[uuid.UUID] = Query(None),
    priority: Optional[str] = Query(None),
    include_archived: bool = Query(False),
    format: str = Query("csv", pattern="^(csv|excel)$"),
):
    """Export tasks with optional filters."""
    q = (
        select(Task)
        .options(
            selectinload(Task.assignee),
            selectinload(Task.creator),
            selectinload(Task.column),
            selectinload(Task.labels),
        )
    )

    if not include_archived:
        q = q.where(Task.is_archived == False)

    # Role-based scoping
    if current_user.role == "superadmin":
        pass
    elif current_user.team_id is not None:
        from sqlalchemy import or_
        Assignee = User.__table__.alias("assignee_user")
        q = q.outerjoin(Assignee, Task.assignee_id == Assignee.c.id).where(
            or_(
                Assignee.c.team_id == current_user.team_id,
                Task.assignee_id.is_(None),
            )
        )
    else:
        q = q.where(Task.assignee_id == current_user.id)

    if board_id:
        q = q.join(KanbanColumn, Task.column_id == KanbanColumn.id).where(
            KanbanColumn.board_id == board_id
        )
    if column_id:
        q = q.where(Task.column_id == column_id)
    if assignee_id:
        q = q.where(Task.assignee_id == assignee_id)
    if priority:
        q = q.where(Task.priority == priority)

    q = q.order_by(Task.created_at.desc())
    result = await db.execute(q)
    tasks = list(result.scalars().all())

    headers = ["Title", "Column", "Assignee", "Priority", "Due Date", "Status", "Labels", "Jira Ticket", "Created At"]
    rows = [
        [
            _csv_safe(task.title),
            _csv_safe(task.column.name if task.column else ""),
            _csv_safe(task.assignee.full_name if task.assignee else ""),
            task.priority,
            task.due_date.isoformat() if task.due_date else "",
            "Archived" if task.is_archived else "Active",
            _csv_safe(", ".join(label.name for label in task.labels) if task.labels else ""),
            task.jira_ticket or "",
            task.created_at.strftime("%Y-%m-%d %H:%M") if task.created_at else "",
        ]
        for task in tasks
    ]

    if format == "excel":
        return _excel_response("tasks.xlsx", rows, headers)

    return _csv_response("tasks.csv", rows, headers)


# ─── User Activity Export ─────────────────────────────────────────────────────

@router.get("/user-activity")
async def export_user_activity(
    current_user: Annotated[User, Depends(get_current_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    user_id: Optional[uuid.UUID] = Query(None),
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    format: str = Query("csv", pattern="^(csv|excel)$"),
):
    """Export per-user daily activity summary."""
    from app.models.worklog import WorkLog
    from sqlalchemy import func

    # Determine which users to export
    if current_user.role == "superadmin":
        target_user_id = user_id
    elif current_user.role == "team_manager":
        target_user_id = user_id  # must be in same team, validated below
    else:
        target_user_id = current_user.id

    q = (
        select(WorkLog)
        .options(selectinload(WorkLog.user), selectinload(WorkLog.work_type))
    )

    if current_user.role == "team_manager":
        # Always scope a manager to their own team, then optionally narrow to a user.
        from app.models.user import User as UserModel
        q = q.join(UserModel, WorkLog.user_id == UserModel.id).where(
            UserModel.team_id == current_user.team_id
        )
        if target_user_id:
            q = q.where(WorkLog.user_id == target_user_id)
    elif target_user_id:
        q = q.where(WorkLog.user_id == target_user_id)

    if date_from:
        q = q.where(WorkLog.log_date >= date_from)
    if date_to:
        q = q.where(WorkLog.log_date <= date_to)

    q = q.order_by(WorkLog.log_date, WorkLog.user_id)
    result = await db.execute(q)
    logs = list(result.scalars().all())

    headers = ["Date", "User", "Email", "Work Type", "Hours", "Description"]
    rows = [
        [
            log.log_date.isoformat() if log.log_date else "",
            log.user.full_name if log.user else "",
            log.user.email if log.user else "",
            log.work_type.name if log.work_type else "",
            str(log.duration_hours),
            log.description or "",
        ]
        for log in logs
    ]

    if format == "excel":
        return _excel_response("user_activity.xlsx", rows, headers)

    return _csv_response("user_activity.csv", rows, headers)


# ─── Excel helper ─────────────────────────────────────────────────────────────

def _excel_response(filename: str, rows: list[list], headers: list[str]) -> StreamingResponse:
    """Return an Excel file using openpyxl."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Export"

        # Header row styling
        header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Data rows
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-width columns
        for col_idx, _ in enumerate(headers, start=1):
            letter = get_column_letter(col_idx)
            max_len = max(
                len(str(ws.cell(row=r, column=col_idx).value or ""))
                for r in range(1, len(rows) + 2)
            )
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 50)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return StreamingResponse(
            iter([buffer.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except ImportError:
        # Fall back to CSV if openpyxl not installed
        return _csv_response(filename.replace(".xlsx", ".csv"), rows, headers)
