import io
import uuid
from datetime import datetime, timezone
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy import select, or_, func as sa_func, update as sa_update
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.inventory import InventoryItem, InventoryEmailSchedule, InventoryGroup
from app.core.security import encrypt_field, decrypt_field
from app.core.exceptions import NotFoundError, ValidationError, ConflictError
from app.config import settings
from app.services.report_schedule_service import compute_next_run


# Mapping: plaintext field name → (encrypted column name, settings key attr)
_ENCRYPTED_FIELDS = {
    "password": "password_encrypted",
    "ssh_key": "ssh_key_encrypted",
    "access_key_id": "access_key_id_encrypted",
    "secret_access_key": "secret_access_key_encrypted",
}

_TYPE_ORDER = ["server", "database", "email_account", "cloud_account", "generic"]

_TYPE_TITLE = {
    "server": "Sunucular",
    "database": "Veritabanları",
    "email_account": "E-posta Hesapları",
    "cloud_account": "Bulut Hesapları",
    "generic": "Genel",
}

_TYPE_COLOR = {
    "server": "2563EB",
    "database": "7C3AED",
    "email_account": "D97706",
    "cloud_account": "EA580C",
    "generic": "4B5563",
}

_COMMON_TAIL_HEADERS = ["Grup", "Etiketler", "Notlar", "Aktif", "Oluşturulma"]

_TYPE_HEADERS: dict[str, list[str]] = {
    "server": [
        "ID", "Ad", "Açıklama", "Hostname", "IP Adresi", "Port",
        "Kullanıcı Adı", "Şifre", "SSH Anahtarı", "İşletim Sistemi",
    ] + _COMMON_TAIL_HEADERS,
    "database": [
        "ID", "Ad", "Açıklama", "Hostname", "IP Adresi", "Port",
        "Kullanıcı Adı", "Şifre", "Veritabanı Adı", "Veritabanı Türü",
    ] + _COMMON_TAIL_HEADERS,
    "email_account": [
        "ID", "Ad", "Açıklama", "E-posta Adresi", "Kullanıcı Adı", "Şifre",
        "SMTP Sunucusu", "SMTP Port", "IMAP Sunucusu", "IMAP Port",
    ] + _COMMON_TAIL_HEADERS,
    "cloud_account": [
        "ID", "Ad", "Açıklama", "Sağlayıcı", "Hesap ID", "Erişim Anahtarı", "Bölge",
    ] + _COMMON_TAIL_HEADERS,
    "generic": [
        "ID", "Ad", "Açıklama", "URL",
    ] + _COMMON_TAIL_HEADERS,
}


class InventoryService:
    def __init__(self, db: AsyncSession):
        self.db = db

    # ─── Item CRUD ────────────────────────────────────────────────────────────

    async def list_items(
        self,
        *,
        item_type: Optional[str] = None,
        search: Optional[str] = None,
        tags: Optional[list[str]] = None,
        is_active: Optional[bool] = None,
        group_id: Optional[uuid.UUID] = None,
        skip: int = 0,
        limit: int = 200,
    ) -> list[InventoryItem]:
        q = select(InventoryItem).options(selectinload(InventoryItem.group))

        if item_type:
            q = q.where(InventoryItem.item_type == item_type)
        if search:
            term = f"%{search}%"
            q = q.where(
                or_(
                    InventoryItem.display_name.ilike(term),
                    InventoryItem.hostname.ilike(term),
                    InventoryItem.ip_address.ilike(term),
                    InventoryItem.email_address.ilike(term),
                    InventoryItem.description.ilike(term),
                )
            )
        if tags:
            for tag in tags:
                q = q.where(InventoryItem.tags.contains([tag]))
        if is_active is not None:
            q = q.where(InventoryItem.is_active == is_active)
        if group_id is not None:
            q = q.where(InventoryItem.group_id == group_id)

        q = q.order_by(InventoryItem.display_name).offset(skip).limit(limit)
        result = await self.db.execute(q)
        return list(result.scalars().all())

    async def get_item(self, item_id: uuid.UUID) -> InventoryItem:
        result = await self.db.execute(
            select(InventoryItem)
            .options(selectinload(InventoryItem.group))
            .where(InventoryItem.id == item_id)
        )
        item = result.scalar_one_or_none()
        if not item:
            raise NotFoundError("Envanter öğesi")
        return item

    async def create_item(
        self, data: dict, created_by: uuid.UUID
    ) -> InventoryItem:
        """Create an inventory item, encrypting sensitive fields."""
        kwargs = {k: v for k, v in data.items() if k not in _ENCRYPTED_FIELDS and v is not None}
        kwargs["created_by"] = created_by
        kwargs["updated_by"] = created_by

        item = InventoryItem(**kwargs)

        key = settings.INVENTORY_ENCRYPTION_KEY
        for plain_field, enc_col in _ENCRYPTED_FIELDS.items():
            plaintext = data.get(plain_field)
            if plaintext:
                if not key:
                    raise ValidationError("INVENTORY_ENCRYPTION_KEY yapılandırılmamış.")
                setattr(item, enc_col, encrypt_field(plaintext, key))

        self.db.add(item)
        await self.db.flush()
        # Re-query to load the group relationship via selectinload.
        return await self.get_item(item.id)

    async def update_item(
        self, item_id: uuid.UUID, data: dict, updated_by: uuid.UUID
    ) -> InventoryItem:
        item = await self.get_item(item_id)

        key = settings.INVENTORY_ENCRYPTION_KEY
        for field, value in data.items():
            if field == "group_id":
                # Allow explicit None to remove item from its group.
                setattr(item, "group_id", value)
                continue
            if value is None:
                continue
            if field in _ENCRYPTED_FIELDS:
                if not key:
                    raise ValidationError("INVENTORY_ENCRYPTION_KEY yapılandırılmamış.")
                setattr(item, _ENCRYPTED_FIELDS[field], encrypt_field(value, key))
            else:
                setattr(item, field, value)

        item.updated_by = updated_by
        await self.db.flush()
        return await self.get_item(item_id)

    # ─── Group CRUD ───────────────────────────────────────────────────────────

    async def list_groups(self) -> list[dict]:
        groups_result = await self.db.execute(
            select(InventoryGroup).order_by(InventoryGroup.name)
        )
        groups = list(groups_result.scalars().all())

        # Fetch item counts in a single query
        counts_result = await self.db.execute(
            select(InventoryItem.group_id, sa_func.count(InventoryItem.id).label("cnt"))
            .where(InventoryItem.group_id.isnot(None))
            .group_by(InventoryItem.group_id)
        )
        count_map: dict[uuid.UUID, int] = {row.group_id: row.cnt for row in counts_result}

        return [
            {
                "id": g.id,
                "name": g.name,
                "description": g.description,
                "group_type": g.group_type,
                "color": g.color,
                "item_count": count_map.get(g.id, 0),
                "created_at": g.created_at,
                "updated_at": g.updated_at,
            }
            for g in groups
        ]

    async def get_group(self, group_id: uuid.UUID) -> InventoryGroup:
        result = await self.db.execute(
            select(InventoryGroup).where(InventoryGroup.id == group_id)
        )
        group = result.scalar_one_or_none()
        if not group:
            raise NotFoundError("Grup")
        return group

    async def create_group(
        self, data: "InventoryGroupCreate", created_by: uuid.UUID
    ) -> dict:
        existing = await self.db.execute(
            select(InventoryGroup).where(InventoryGroup.name == data.name)
        )
        if existing.scalar_one_or_none():
            raise ConflictError("Bu isimde bir grup zaten mevcut.")

        group = InventoryGroup(
            name=data.name,
            description=data.description,
            group_type=data.group_type,
            color=data.color,
            created_by=created_by,
        )
        self.db.add(group)
        await self.db.flush()
        await self.db.refresh(group)
        return {"id": group.id, "name": group.name, "description": group.description,
                "group_type": group.group_type, "color": group.color, "item_count": 0,
                "created_at": group.created_at, "updated_at": group.updated_at}

    async def update_group(
        self, group_id: uuid.UUID, data: "InventoryGroupUpdate"
    ) -> dict:
        group = await self.get_group(group_id)

        if data.name is not None and data.name != group.name:
            existing = await self.db.execute(
                select(InventoryGroup).where(InventoryGroup.name == data.name)
            )
            if existing.scalar_one_or_none():
                raise ConflictError("Bu isimde bir grup zaten mevcut.")
            group.name = data.name
        if data.description is not None:
            group.description = data.description
        if data.group_type is not None:
            group.group_type = data.group_type
        if data.color is not None:
            group.color = data.color

        await self.db.flush()
        await self.db.refresh(group)

        count_result = await self.db.execute(
            select(sa_func.count(InventoryItem.id)).where(InventoryItem.group_id == group_id)
        )
        item_count = count_result.scalar_one()
        return {"id": group.id, "name": group.name, "description": group.description,
                "group_type": group.group_type, "color": group.color, "item_count": item_count,
                "created_at": group.created_at, "updated_at": group.updated_at}

    async def delete_group(self, group_id: uuid.UUID) -> None:
        group = await self.get_group(group_id)
        await self.db.delete(group)
        await self.db.flush()

    async def assign_items_to_group(
        self, group_id: Optional[uuid.UUID], item_ids: list[uuid.UUID]
    ) -> int:
        if group_id is not None:
            await self.get_group(group_id)  # raises NotFoundError if missing

        result = await self.db.execute(
            sa_update(InventoryItem)
            .where(InventoryItem.id.in_(item_ids))
            .values(group_id=group_id)
        )
        await self.db.flush()
        return result.rowcount

    async def delete_item(self, item_id: uuid.UUID) -> None:
        item = await self.get_item(item_id)
        await self.db.delete(item)
        await self.db.flush()

    async def reveal_field(self, item_id: uuid.UUID, field_name: str) -> str:
        """Decrypt and return a single sensitive field value."""
        if field_name not in _ENCRYPTED_FIELDS:
            raise ValidationError(f"Geçersiz alan: {field_name}")

        item = await self.get_item(item_id)
        enc_col = _ENCRYPTED_FIELDS[field_name]
        encrypted = getattr(item, enc_col)
        if not encrypted:
            raise ValidationError("Bu alan için kayıtlı veri bulunamadı.")

        key = settings.INVENTORY_ENCRYPTION_KEY
        if not key:
            raise ValidationError("INVENTORY_ENCRYPTION_KEY yapılandırılmamış.")
        return decrypt_field(encrypted, key)

    # ─── Export ───────────────────────────────────────────────────────────────

    async def export_excel(self, item_type: Optional[str] = None) -> bytes:
        """Generate a grouped Excel file — one sheet per item type + summary sheet."""
        from collections import defaultdict
        items = await self.list_items(item_type=item_type, limit=10000)

        by_type: dict[str, list] = defaultdict(list)
        for item in items:
            by_type[item.item_type].append(item)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # remove default empty sheet

        # ── Summary sheet ────────────────────────────────────────────────────
        ws_sum = wb.create_sheet("Özet")
        ws_sum["A1"] = "Envanter Özeti"
        ws_sum["A1"].font = Font(bold=True, size=14, color="1E293B")
        ws_sum.merge_cells("A1:E1")

        ws_sum["A2"] = f"Oluşturulma: {datetime.now(timezone.utc).strftime('%d.%m.%Y %H:%M')} UTC"
        ws_sum["A2"].font = Font(italic=True, color="6B7280", size=10)
        ws_sum.merge_cells("A2:E2")

        dark_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        dark_font = Font(color="FFFFFF", bold=True)
        center = Alignment(horizontal="center")

        for col_idx, hdr in enumerate(["Tür", "Toplam", "Aktif", "Pasif", "Sayfa Adı"], start=1):
            cell = ws_sum.cell(row=4, column=col_idx, value=hdr)
            cell.fill = dark_fill
            cell.font = dark_font
            cell.alignment = center

        grand = [0, 0, 0]
        types_written = [tp for tp in _TYPE_ORDER if by_type.get(tp) and (item_type is None or tp == item_type)]

        for row_offset, tp in enumerate(types_written, start=1):
            type_items = by_type[tp]
            active = sum(1 for i in type_items if i.is_active)
            inactive = len(type_items) - active
            grand[0] += len(type_items); grand[1] += active; grand[2] += inactive

            tp_fill = PatternFill(start_color=_TYPE_COLOR[tp], end_color=_TYPE_COLOR[tp], fill_type="solid")
            r = 4 + row_offset
            c1 = ws_sum.cell(row=r, column=1, value=_TYPE_TITLE[tp])
            c1.fill = tp_fill; c1.font = Font(color="FFFFFF", bold=True)
            ws_sum.cell(row=r, column=2, value=len(type_items)).alignment = center
            ws_sum.cell(row=r, column=3, value=active).font = Font(color="16A34A", bold=True)
            c4 = ws_sum.cell(row=r, column=4, value=inactive)
            c4.font = Font(color="DC2626", bold=True) if inactive else Font(color="6B7280")
            ws_sum.cell(row=r, column=5, value=_TYPE_TITLE[tp]).alignment = center

        total_row = 4 + len(types_written) + 2
        ws_sum.cell(row=total_row, column=1, value="TOPLAM").font = Font(bold=True, size=11)
        ws_sum.cell(row=total_row, column=2, value=grand[0]).font = Font(bold=True, size=11)
        c3 = ws_sum.cell(row=total_row, column=3, value=grand[1])
        c3.font = Font(bold=True, color="16A34A", size=11)
        c4 = ws_sum.cell(row=total_row, column=4, value=grand[2])
        c4.font = Font(bold=True, color="DC2626" if grand[2] else "6B7280", size=11)

        ws_sum.column_dimensions["A"].width = 26
        for col in "BCDE":
            ws_sum.column_dimensions[col].width = 14

        # ── Per-type sheets ──────────────────────────────────────────────────
        for tp in _TYPE_ORDER:
            type_items = by_type.get(tp, [])
            if not type_items or (item_type and tp != item_type):
                continue

            ws = wb.create_sheet(_TYPE_TITLE[tp])
            try:
                ws.sheet_properties.tabColor = _TYPE_COLOR[tp]
            except Exception:
                pass

            tp_fill = PatternFill(start_color=_TYPE_COLOR[tp], end_color=_TYPE_COLOR[tp], fill_type="solid")
            hdr_font = Font(color="FFFFFF", bold=True)

            for col_idx, hdr in enumerate(_TYPE_HEADERS[tp], start=1):
                cell = ws.cell(row=1, column=col_idx, value=hdr)
                cell.fill = tp_fill
                cell.font = hdr_font
                cell.alignment = center

            for row_idx, item in enumerate(type_items, start=2):
                for col_idx, value in enumerate(_item_to_typed_row(item), start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    async def export_csv(self, item_type: Optional[str] = None) -> str:
        """Generate grouped CSV — each item type has its own header row, sections separated by blank lines."""
        import csv as _csv
        from collections import defaultdict
        items = await self.list_items(item_type=item_type, limit=10000)

        by_type: dict[str, list] = defaultdict(list)
        for item in items:
            by_type[item.item_type].append(item)

        output = io.StringIO()
        writer = _csv.writer(output)

        types_to_write = [
            tp for tp in _TYPE_ORDER
            if by_type.get(tp) and (item_type is None or tp == item_type)
        ]

        for section_idx, tp in enumerate(types_to_write):
            type_items = by_type[tp]
            if section_idx > 0:
                writer.writerow([])  # blank line between sections

            writer.writerow([f"=== {_TYPE_TITLE[tp]} ({len(type_items)}) ==="])
            writer.writerow(_TYPE_HEADERS[tp])
            for item in type_items:
                writer.writerow(_item_to_typed_row(item))

        return output.getvalue()

    # ─── Email Schedules ─────────────────────────────────────────────────────

    async def list_schedules(self) -> list[InventoryEmailSchedule]:
        result = await self.db.execute(
            select(InventoryEmailSchedule).order_by(InventoryEmailSchedule.created_at)
        )
        return list(result.scalars().all())

    async def get_schedule(self, schedule_id: uuid.UUID) -> InventoryEmailSchedule:
        result = await self.db.execute(
            select(InventoryEmailSchedule).where(InventoryEmailSchedule.id == schedule_id)
        )
        sch = result.scalar_one_or_none()
        if not sch:
            raise NotFoundError("Zamanlama")
        return sch

    async def create_schedule(
        self, data: dict, created_by: uuid.UUID
    ) -> InventoryEmailSchedule:
        sch = InventoryEmailSchedule(
            **data,
            created_by=created_by,
            next_run_at=compute_next_run(
                data["frequency"],
                data.get("day_of_week"),
                data.get("day_of_month"),
                data.get("hour", 8),
            ),
        )
        self.db.add(sch)
        await self.db.flush()
        await self.db.refresh(sch)
        return sch

    async def update_schedule(
        self, schedule_id: uuid.UUID, data: dict
    ) -> InventoryEmailSchedule:
        sch = await self.get_schedule(schedule_id)
        for field, value in data.items():
            if value is not None:
                setattr(sch, field, value)
        # Recompute next run
        sch.next_run_at = compute_next_run(
            sch.frequency, sch.day_of_week, sch.day_of_month, sch.hour
        )
        await self.db.flush()
        await self.db.refresh(sch)
        return sch

    async def delete_schedule(self, schedule_id: uuid.UUID) -> None:
        sch = await self.get_schedule(schedule_id)
        await self.db.delete(sch)
        await self.db.flush()

    async def send_now(self, schedule_id: uuid.UUID) -> int:
        """Manually trigger an inventory email schedule."""
        sch = await self.get_schedule(schedule_id)
        sent = await _send_inventory_email(self.db, sch)
        sch.last_run_at = datetime.now(timezone.utc)
        sch.next_run_at = compute_next_run(
            sch.frequency, sch.day_of_week, sch.day_of_month, sch.hour
        )
        await self.db.flush()
        return sent


# ─── Helpers ─────────────────────────────────────────────────────────────────

def _item_to_typed_row(item: InventoryItem) -> list:
    """Return a type-specific row matching _TYPE_HEADERS[item.item_type]."""
    group_name = item.group.name if item.group else ""
    tags = ", ".join(item.tags) if item.tags else ""
    tail = [group_name, tags, item.notes or "", "Evet" if item.is_active else "Hayır",
            item.created_at.strftime("%Y-%m-%d %H:%M") if item.created_at else ""]

    t = item.item_type
    if t == "server":
        return [
            str(item.id), item.display_name, item.description or "",
            item.hostname or "", item.ip_address or "", item.port or "",
            item.username or "",
            "Var" if item.password_encrypted else "Yok",
            "Var" if item.ssh_key_encrypted else "Yok",
            item.operating_system or "",
        ] + tail
    elif t == "database":
        return [
            str(item.id), item.display_name, item.description or "",
            item.hostname or "", item.ip_address or "", item.port or "",
            item.username or "",
            "Var" if item.password_encrypted else "Yok",
            item.database_name or "", item.database_type or "",
        ] + tail
    elif t == "email_account":
        return [
            str(item.id), item.display_name, item.description or "",
            item.email_address or "", item.username or "",
            "Var" if item.password_encrypted else "Yok",
            item.smtp_host or "", item.smtp_port or "",
            item.imap_host or "", item.imap_port or "",
        ] + tail
    elif t == "cloud_account":
        return [
            str(item.id), item.display_name, item.description or "",
            item.provider or "", item.account_id or "",
            "Var" if item.access_key_id_encrypted else "Yok",
            item.region or "",
        ] + tail
    else:  # generic
        return [
            str(item.id), item.display_name, item.description or "",
            item.url or "",
        ] + tail


async def _send_inventory_email(db: AsyncSession, schedule: InventoryEmailSchedule) -> int:
    """Generate Excel and send to all schedule recipients. Returns count of sent emails."""
    from app.models.email_config import SmtpConfig

    smtp_result = await db.execute(
        select(SmtpConfig).where(SmtpConfig.is_active == True).limit(1)
    )
    smtp = smtp_result.scalar_one_or_none()

    if not smtp or not schedule.recipient_emails:
        return 0

    # Generate inventory Excel
    svc = InventoryService(db)
    excel_bytes = await svc.export_excel()

    sent = 0
    try:
        import smtplib
        from email.mime.multipart import MIMEMultipart
        from email.mime.text import MIMEText
        from email.mime.base import MIMEBase
        from email import encoders
        from app.core.security import decrypt_field as _dec
        from app.config import settings as _s

        smtp_password = _dec(smtp.password_encrypted, _s.SMTP_ENCRYPTION_KEY)

        for recipient in schedule.recipient_emails:
            try:
                msg = MIMEMultipart()
                msg["From"] = f"{smtp.from_name} <{smtp.from_email}>"
                msg["To"] = recipient
                msg["Subject"] = f"Envanter Raporu — {schedule.name}"

                body = "Ekte envanter raporu yer almaktadır.\n"
                msg.attach(MIMEText(body, "plain", "utf-8"))

                part = MIMEBase("application", "octet-stream")
                part.set_payload(excel_bytes)
                encoders.encode_base64(part)
                today_str = datetime.now(timezone.utc).strftime("%Y%m%d")
                part.add_header(
                    "Content-Disposition",
                    f'attachment; filename="inventory_{today_str}.xlsx"',
                )
                msg.attach(part)

                import ssl as _ssl
                _ctx = _ssl.create_default_context()
                if getattr(smtp, 'use_ssl', False):
                    server = smtplib.SMTP_SSL(smtp.host, smtp.port, context=_ctx, timeout=15)
                elif smtp.use_tls:
                    server = smtplib.SMTP(smtp.host, smtp.port, timeout=15)
                    server.ehlo()
                    server.starttls(context=_ctx)
                    server.ehlo()
                else:
                    server = smtplib.SMTP(smtp.host, smtp.port, timeout=15)
                    server.ehlo()

                server.login(smtp.username, smtp_password)
                server.sendmail(smtp.from_email, recipient, msg.as_string())
                server.quit()
                sent += 1
            except Exception:
                pass
    except Exception:
        pass

    return sent


async def run_due_inventory_schedules(db: AsyncSession) -> None:
    """Called by Celery Beat hourly — sends emails for due schedules."""
    from datetime import timezone as _tz
    now = datetime.now(_tz.utc)
    result = await db.execute(
        select(InventoryEmailSchedule).where(
            InventoryEmailSchedule.is_active == True,
            InventoryEmailSchedule.next_run_at <= now,
        )
    )
    schedules = list(result.scalars().all())

    for sch in schedules:
        try:
            await _send_inventory_email(db, sch)
            sch.last_run_at = now
            sch.next_run_at = compute_next_run(
                sch.frequency, sch.day_of_week, sch.day_of_month, sch.hour
            )
        except Exception:
            pass

    if schedules:
        await db.flush()
